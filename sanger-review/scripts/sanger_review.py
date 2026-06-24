#!/usr/bin/env python3
from __future__ import annotations
import argparse,csv,html,re,struct
from collections import Counter,defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Dict,List,Optional,Tuple
import xml.etree.ElementTree as ET
from Bio import SeqIO
from Bio.Align import PairwiseAligner
from Bio.Seq import Seq
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Font,PatternFill
from openpyxl.utils import get_column_letter
try:
    import yaml
except Exception:
    yaml=None
try:
    import matplotlib; matplotlib.use('Agg'); import matplotlib.pyplot as plt
except Exception:
    plt=None
DNA=set('ACGT'); REF_EXTS={'.dna','.fa','.fasta','.fna','.gb','.gbk','.genbank'}
@dataclass
class Ref:
    sample:str; path:Path; seq:str; label:str; start:int; end:int; index:Dict[str,List[int]]
@dataclass
class Read:
    path:Path; sample:str; clone:str; primer:str; seq:str; qual:List[int]; result:str=''; note:str=''
@dataclass
class Call:
    read:Read; orient:str; offset:Optional[int]; seed:int; aligned:int; highq:int; mismatches:List[dict]; indels:List[dict]; bases:Dict[int,str]
@dataclass
class Target:
    sample:str; name:str; aa:Optional[int]; positions:List[int]; expected:str

def load_cfg(p):
    if not p: return {}
    if yaml is None: raise RuntimeError('PyYAML is required for --config')
    return yaml.safe_load(Path(p).read_text(encoding='utf-8')) or {}
def val(args,cfg,name,default=None):
    v=getattr(args,name,None)
    return v if v is not None else cfg.get(name.replace('_','-'),cfg.get(name,default))
def packets(b):
    i=0
    while i+5<=len(b):
        t=b[i]; n=struct.unpack('>I',b[i+1:i+5])[0]; s=i+5; e=s+n
        if e>len(b): break
        yield t,b[s:e]; i=e
def kindex(seq,k=18):
    d=defaultdict(list)
    for i in range(max(0,len(seq)-k+1)):
        x=seq[i:i+k]
        if set(x)<=DNA: d[x].append(i)
    return dict(d)
def parse_region(x):
    if not x: return None
    m=re.match(r'^\s*(\d+)\s*[-:]\s*(\d+)\s*$',str(x))
    if not m: raise ValueError(f'invalid region {x!r}')
    a,b=int(m.group(1)),int(m.group(2)); return (min(a,b),max(a,b))
def id_from_name(name,regex,stem):
    if not regex: return stem
    m=re.search(regex,name)
    if not m: raise ValueError(f'cannot parse id from {name!r}')
    return m.groupdict().get('sample') or m.group(1)
def feature_from_snap(xml,feature,region):
    if region: return region
    if not feature: raise ValueError('use --feature or --region for SnapGene')
    root=ET.fromstring(xml)
    for f in root.findall('Feature'):
        if f.attrib.get('name')==feature or f.attrib.get('type')==feature:
            seg=f.find('Segment')
            if seg is not None and 'range' in seg.attrib:
                a,b=seg.attrib['range'].split('-'); return int(a),int(b)
    raise ValueError(f'feature {feature!r} not found')
def read_snap(p,regex,feature,region):
    seq=None; xml=None
    for t,c in packets(p.read_bytes()):
        if t==0 and c: seq=c[1:].decode('ascii')
        elif t==10 and c.startswith(b'<?xml'): xml=c.decode('utf-8',errors='replace')
    if not seq: raise ValueError(f'no DNA packet in {p}')
    a,b=feature_from_snap(xml,feature,region) if xml else region
    if not a: raise ValueError(f'no feature XML in {p}; use --region')
    s=seq.upper(); return Ref(id_from_name(p.name,regex,p.stem),p,s,p.stem,a,b,kindex(s))
def read_seqref(p,regex,feature,region):
    fmt='fasta' if p.suffix.lower() in {'.fa','.fasta','.fna'} else 'genbank'; rec=SeqIO.read(str(p),fmt); seq=str(rec.seq).upper(); a,b=1,len(seq)
    if region: a,b=region
    elif fmt=='genbank' and feature:
        ok=False
        for ft in rec.features:
            names=[]
            for key in ('gene','label','note','product'): names+=ft.qualifiers.get(key,[])
            if ft.type==feature or feature in names: a=int(ft.location.start)+1; b=int(ft.location.end); ok=True; break
        if not ok: raise ValueError(f'feature {feature!r} not found in {p}')
    elif fmt=='fasta' and feature: raise ValueError('FASTA has no features; use --region')
    return Ref(id_from_name(p.name,regex,p.stem),p,seq,p.stem,a,b,kindex(seq))
def load_refs(path,regex,feature,region):
    path=Path(path); files=[path] if path.is_file() else sorted(x for x in path.rglob('*') if x.suffix.lower() in REF_EXTS)
    refs={}
    for p in files:
        r=read_snap(p,regex,feature,region) if p.suffix.lower()=='.dna' else read_seqref(p,regex,feature,region)
        if r.sample in refs: raise ValueError(f'duplicate reference id {r.sample}')
        refs[r.sample]=r
    if not refs: raise ValueError('no reference files found')
    return refs

def table(path):
    if not path or not Path(path).exists(): return []
    path=Path(path)
    if path.suffix.lower() in {'.xlsx','.xlsm'}:
        wb=load_workbook(path,read_only=True,data_only=True); ws=wb.active; rows=ws.iter_rows(values_only=True); hdr=[str(x or '').strip() for x in next(rows)]
        return [dict(zip(hdr,r)) for r in rows]
    with path.open(newline='',encoding='utf-8-sig') as h: return list(csv.DictReader(h))
def metadata(path):
    d={}
    for r in table(path):
        s=r.get('sample') or r.get('sample_name') or r.get('Sample') or r.get('Sample Name') or r.get('????')
        pr=r.get('primer') or r.get('Primer') or r.get('????') or 'read'
        if s: d[(str(s),str(pr))]={'result':str(r.get('result') or r.get('Result') or r.get('??') or ''),'note':str(r.get('note') or r.get('remark') or r.get('Remark') or r.get('????') or '')}
    return d
def read_id(p,regex):
    m=re.search(regex,p.name)
    if not m: raise ValueError(f'cannot parse read filename {p.name!r}')
    gd=m.groupdict(); return str(gd.get('sample') or m.group(1)),str(gd.get('clone') or gd.get('replicate') or '1'),str(gd.get('primer') or 'read')
def read_ab1(p,regex,meta):
    s,c,pr=read_id(p,regex); rec=SeqIO.read(str(p),'abi'); seq=str(rec.seq).upper(); q=list(rec.letter_annotations.get('phred_quality',[]))
    if len(q)!=len(seq): q=[0]*len(seq)
    md=meta.get((s,pr),{})
    return Read(p,s,c,pr,seq,q,md.get('result',''),md.get('note',''))
def read_files(path):
    path=Path(path); return [path] if path.is_file() else sorted(path.rglob('*.ab1'))
def orient(read,rev):
    return (str(Seq(read.seq).reverse_complement()),list(reversed(read.qual)),'-') if rev else (read.seq,read.qual,'+')
def offset(seq,qual,ref,k=18):
    cnt=Counter()
    for mq in (18,12,0):
        cnt.clear(); step=4 if mq else 2
        for i in range(0,max(len(seq)-k+1,0),step):
            x=seq[i:i+k]
            if set(x)<=DNA and all(q>=mq for q in qual[i:i+k]):
                for rp in ref.index.get(x,[]): cnt[rp-i]+=1
        if cnt:
            off,score=cnt.most_common(1)[0]
            if score>=(2 if mq else 1): return off,score
    return None,0
def gapped(target,query,aln):
    tb,qb=aln.aligned; rp=[]; qp=[]; pt=pq=0
    for (ts,te),(qs,qe) in zip(tb,qb):
        if ts>pt: rp.append(target[pt:ts]); qp.append('-'*(ts-pt))
        if qs>pq: rp.append('-'*(qs-pq)); qp.append(query[pq:qs])
        rp.append(target[ts:te]); qp.append(query[qs:qe]); pt,pq=te,qe
    if pt<len(target): rp.append(target[pt:]); qp.append('-'*(len(target)-pt))
    if pq<len(query): rp.append('-'*(len(query)-pq)); qp.append(query[pq:])
    return ''.join(rp),''.join(qp)
def prob(kind,pos,fpos,exp,obs,q,read): return {'kind':kind,'position':pos,'feature_position':fpos,'expected':exp,'observed':obs,'quality':q,'read':read}
def compare(read,seq,qual,ori,off,seed,ref,minq,trim):
    if off is None: return Call(read,ori,None,seed,0,0,[],[],{})
    rs=max(0,off); qs=max(0,-off); ov=min(len(ref.seq)-rs,len(seq)-qs)
    if ov<=0: return Call(read,ori,None,seed,0,0,[],[],{})
    target=ref.seq[rs:rs+ov]; query=seq[qs:qs+ov]; qqual=qual[qs:qs+ov]
    al=PairwiseAligner(); al.mode='global'; al.match_score=2; al.mismatch_score=-1; al.open_gap_score=-5; al.extend_gap_score=-0.5
    ra,qa=gapped(target,query,al.align(target,query)[0])
    mm=[]; ind=[]; bases={}; aligned=highq=0; ri=qi=0; last=rs
    for rb,qb in zip(ra,qa):
        if rb!='-': ri+=1; pos=rs+ri; last=pos
        else: pos=last
        if qb!='-': qi+=1
        if not(ref.start<=pos<=ref.end): continue
        fpos=pos-ref.start+1
        if rb!='-' and qb!='-':
            oq=qs+qi-1
            if oq<trim or oq>=len(seq)-trim: continue
            aligned+=1; q=qqual[qi-1] if qi-1<len(qqual) else 0
            if q>=minq and qb in DNA:
                highq+=1; bases[pos]=qb
                if qb!=rb: mm.append(prob('mismatch',pos,fpos,rb,qb,q,read.path.name))
        elif rb!='-' and qb=='-':
            nq=qs+qi
            if nq<trim or nq>=len(seq)-trim: continue
            ind.append(prob('deletion',pos,fpos,rb,'','',read.path.name))
        elif rb=='-' and qb!='-':
            oq=qs+qi-1
            if oq<trim or oq>=len(seq)-trim: continue
            q=qqual[qi-1] if qi-1<len(qqual) else 0
            if q>=minq: ind.append(prob('insertion',pos,fpos,'-',qb,q,read.path.name))
    return Call(read,ori,off,seed,aligned,highq,mm,ind,bases)
def align(read,ref,minq,trim):
    calls=[]
    for rev in (False,True):
        s,q,o=orient(read,rev); off,seed=offset(s,q,ref); calls.append(compare(read,s,q,o,off,seed,ref,minq,trim))
    return max(calls,key=lambda c:(c.highq,c.aligned,c.seed,-len(c.mismatches)))
def select(calls):
    g=defaultdict(list)
    for c in calls: g[c.read.primer].append(c)
    return [max(v,key=lambda c:(c.highq,c.aligned,c.seed,-len(c.mismatches),-len(c.indels))) for v in g.values()]

def targets_from_label(ref):
    out=[]
    for m in re.finditer(r'\b([A-Z*])(\d+)([A-Z*])\b',ref.label):
        aa=int(m.group(2)); st=ref.start+(aa-1)*3; pos=[st,st+1,st+2]
        if all(1<=x<=len(ref.seq) for x in pos): out.append(Target(ref.sample,m.group(0),aa,pos,''.join(ref.seq[x-1] for x in pos)))
    return out
def load_targets(path,refs,from_names):
    d=defaultdict(list)
    if from_names:
        for r in refs.values(): d[r.sample]+=targets_from_label(r)
    for row in table(path):
        s=str(row.get('sample') or row.get('reference') or row.get('mutant') or '').strip()
        if s not in refs: continue
        r=refs[s]; name=str(row.get('name') or row.get('mutation') or row.get('target') or 'target').strip(); exp=str(row.get('expected') or '').strip().upper()
        aa=row.get('aa_position') or row.get('position_aa'); cp=row.get('codon_start') or row.get('map_position') or row.get('position')
        if aa:
            aa=int(aa); st=r.start+(aa-1)*3; pos=[st,st+1,st+2]
        elif cp:
            aa=None; st=int(cp); pos=[st] if len(exp)==1 else [st,st+1,st+2]
        else: continue
        if not exp: exp=''.join(r.seq[x-1] for x in pos)
        d[s].append(Target(s,name,aa,pos,exp))
    return dict(d)
def base_sets(calls):
    d=defaultdict(set)
    for c in calls:
        for p,b in c.bases.items(): d[p].add(b)
    return d
def eval_targets(sample,clone,ref,calls,targets):
    bs=base_sets(calls); rows=[]; ok=miss=bad=0; last=ref.start-1
    for t in targets:
        last=max(last,max(t.positions)); obs=[]; states=[]
        for p in t.positions:
            b=bs.get(p,set()); exp=ref.seq[p-1]
            if not b: obs.append('?'); states.append('missing')
            elif len(b)>1: obs.append('/'.join(sorted(b))); states.append('conflict')
            else:
                x=next(iter(b)); obs.append(x); states.append('match' if x==exp else 'mismatch')
        if all(x=='match' for x in states): status='covered_match'; ok+=1
        elif any(x in ('mismatch','conflict') for x in states): status='problem'; bad+=1
        else: status='missing_or_partial'; miss+=1
        rows.append({'sample':sample,'clone':clone,'target':t.name,'aa_position':t.aa or '','map_positions':'-'.join(map(str,t.positions)),'feature_positions':'-'.join(str(p-ref.start+1) for p in t.positions),'expected':t.expected,'observed':''.join(obs),'status':status})
    return rows,ok,miss,bad,last
def norm_call(x):
    m={'candidate':'Candidate','pick clone':'Candidate','needs review':'Needs review','needs resequencing':'Needs review','review':'Needs review','reject':'No good clone','discard':'No good clone','no good clone':'No good clone'}
    k=(x or '').strip().lower().replace('_',' ')
    if k not in m: raise ValueError(f'unsupported manual_call {x!r}')
    return m[k]
def overrides(path):
    d={}
    for r in table(path):
        s=str(r.get('sample') or r.get('mutant') or '').strip(); c=str(r.get('clone') or '').strip(); call=str(r.get('manual_call') or '').strip()
        if s and c and call: d[(s,c)]={'manual_override':call,'final_call':norm_call(call),'manual_note':str(r.get('note') or r.get('manual_note') or '').strip()}
    return d
def clone_call(mode,tmiss,tbad,mm,conf,ind,down,has_targets):
    if mode=='differences-only': return 'Review','differences-only mode does not assign candidate status','review differences'
    if mode=='strict':
        if tmiss or tbad or mm or conf or ind: return 'Needs review','strict mode requires complete target-region agreement','review or resequence clone'
        return 'Candidate','strict mode found no target-region differences','pick clone'
    if tbad or mm or conf: return 'No good clone','target site problem or internal high-quality mismatch/conflict detected','discard clone'
    if has_targets and tmiss: return 'Needs review','one or more target sites are not fully covered by trusted read bases','review or resequence clone'
    if (ind-down if has_targets else ind)>0: return 'Needs review','base caller suggests indel before or within target span; inspect chromatogram','inspect peaks'
    return 'Candidate','target sites match and no blocking mismatch/conflict was detected','pick clone'
def plot_chrom(read,outdir):
    if plt is None: return None
    rec=SeqIO.read(str(read.path),'abi'); raw=rec.annotations.get('abif_raw',{}); traces=[raw.get(f'DATA{i}') for i in range(9,13)]; ploc=raw.get('PLOC2') or raw.get('PLOC1') or []; bases=(raw.get('PBAS2') or raw.get('PBAS1') or b'').decode('ascii',errors='ignore'); fwo=(raw.get('FWO_1') or b'GATC').decode('ascii',errors='ignore')
    if not all(traces) or not ploc: return None
    outdir.mkdir(parents=True,exist_ok=True); colors={'A':'#2ca02c','C':'#1f77b4','G':'#000000','T':'#d62728'}; fig,ax=plt.subplots(figsize=(16,4),dpi=150)
    for b,t in zip(fwo,traces): ax.plot(t,color=colors.get(b,'#777777'),linewidth=.7,label=b)
    ymax=max(max(t) for t in traces); step=max(1,len(ploc)//80)
    for i in range(0,min(len(ploc),len(bases)),step): ax.text(ploc[i],ymax*1.02,bases[i],fontsize=6,ha='center',color=colors.get(bases[i],'#333333'))
    ax.set_xlim(0,max(len(t) for t in traces)); ax.set_ylim(0,ymax*1.12); ax.set_title(read.path.name); ax.set_xlabel('Trace index'); ax.set_ylabel('Signal'); ax.legend(loc='upper right',ncol=4,fontsize=8); fig.tight_layout()
    p=outdir/(read.path.stem+'.png'); fig.savefig(p); plt.close(fig); return p
def fill(v): return PatternFill('solid',fgColor={'Candidate':'C6EFCE','Needs review':'FFEB9C','Review':'D9EAF7','No good clone':'FFC7CE','covered_match':'C6EFCE','missing_or_partial':'FFEB9C','problem':'FFC7CE'}.get(v,'FFFFFF'))
def workbook(path,sheets):
    wb=Workbook()
    for i,(name,rows) in enumerate(sheets):
        ws=wb.active if i==0 else wb.create_sheet(); ws.title=name; hdr=list(rows[0].keys()) if rows else ['message']; ws.append(hdr)
        for c in ws[1]: c.font=Font(bold=True)
        for r in rows or [{'message':'No rows'}]: ws.append([r.get(h,'') for h in hdr])
        for s in ('summary_call','final_call','auto_call','status'):
            if s in hdr:
                col=hdr.index(s)+1
                for rr in range(2,ws.max_row+1): ws.cell(rr,col).fill=fill(str(ws.cell(rr,col).value))
        for col in range(1,ws.max_column+1):
            w=12
            for rr in range(1,min(ws.max_row,200)+1):
                v=ws.cell(rr,col).value
                if v is not None: w=max(w,min(len(str(v))+2,80))
            ws.column_dimensions[get_column_letter(col)].width=w
        ws.freeze_panes='A2'; ws.auto_filter.ref=ws.dimensions
    wb.save(path)
def html_report(path,summary,clones,problems):
    def tbl(rows,limit=None):
        if not rows: return '<p>No rows.</p>'
        rows=rows[:limit] if limit else rows; hdr=list(rows[0].keys()); out=['<table><thead><tr>']+[f'<th>{html.escape(h)}</th>' for h in hdr]+['</tr></thead><tbody>']
        for r in rows:
            cls=str(r.get('summary_call') or r.get('final_call') or r.get('status') or '').lower().replace(' ','-'); out.append(f"<tr class='{html.escape(cls)}'>"); out += [f'<td>{html.escape(str(r.get(h,"")))}</td>' for h in hdr]; out.append('</tr>')
        out.append('</tbody></table>'); return '\n'.join(out)
    css="body{font-family:Arial,sans-serif;margin:24px;color:#222}table{border-collapse:collapse;width:100%;margin:12px 0 28px}th,td{border:1px solid #ccc;padding:6px 8px;font-size:13px;text-align:left;vertical-align:top}th{background:#f2f2f2}.candidate td{background:#e2f0d9}.needs-review td{background:#fff2cc}.review td{background:#ddebf7}.no-good-clone td{background:#f4cccc}"
    path.write_text('\n'.join(['<!doctype html><html><head><meta charset="utf-8"><title>Sanger Review</title>',f'<style>{css}</style></head><body>','<h1>Sanger Review</h1>','<h2>Summary</h2>',tbl(summary),'<h2>Clone Review</h2>',tbl(clones),'<h2>Problem Sites</h2>',tbl(problems,500),'</body></html>']),encoding='utf-8')

def main(argv=None):
    pa=argparse.ArgumentParser(description='Review Sanger AB1 reads against reference sequences and call candidate clones.')
    pa.add_argument('--config',type=Path); pa.add_argument('--references',type=Path); pa.add_argument('--reads',type=Path); pa.add_argument('--out',type=Path)
    pa.add_argument('--feature'); pa.add_argument('--region'); pa.add_argument('--reference-regex'); pa.add_argument('--sample-regex'); pa.add_argument('--target-sites',type=Path); pa.add_argument('--targets-from-reference-names',action='store_true'); pa.add_argument('--manual-overrides',type=Path); pa.add_argument('--metadata',type=Path)
    pa.add_argument('--quality',type=int); pa.add_argument('--trim-read-ends',type=int); pa.add_argument('--mode',choices=['target-first','strict','differences-only']); pa.add_argument('--chromatograms',choices=['none','problem','all']); pa.add_argument('--prefix')
    a=pa.parse_args(argv); cfg=load_cfg(a.config)
    refs_path=Path(val(a,cfg,'references')); reads_path=Path(val(a,cfg,'reads')); out=Path(val(a,cfg,'out','outputs/sanger_review')); out.mkdir(parents=True,exist_ok=True); chrom=out/'chromatograms'
    feature=val(a,cfg,'feature'); region=parse_region(val(a,cfg,'region')); ref_regex=val(a,cfg,'reference_regex')
    sample_regex=val(a,cfg,'sample_regex',r'(?P<sample>.+?)[-_](?P<clone>\d+)[-_](?P<primer>[^_.]+).*\.ab1$')
    minq=int(val(a,cfg,'quality',20)); trim=int(val(a,cfg,'trim_read_ends',100)); mode=val(a,cfg,'mode','target-first'); chro=val(a,cfg,'chromatograms','problem'); prefix=val(a,cfg,'prefix','sanger_review')
    targets_path=val(a,cfg,'target_sites'); manual_path=val(a,cfg,'manual_overrides'); meta_path=val(a,cfg,'metadata'); from_names=bool(a.targets_from_reference_names or cfg.get('targets_from_reference_names',False) or cfg.get('targets-from-reference-names',False))
    refs=load_refs(refs_path,ref_regex,feature,region); targs=load_targets(Path(targets_path) if targets_path else None,refs,from_names); ovs=overrides(Path(manual_path) if manual_path else (out/'manual_overrides.csv')); meta=metadata(Path(meta_path) if meta_path else None)
    reads=[]
    for p in read_files(reads_path):
        try:
            r=read_ab1(p,sample_regex,meta)
            if r.sample in refs: reads.append(r)
            else: print(f'WARNING: no reference for read sample {r.sample!r}: {r.path.name}')
        except Exception as e: print(f'WARNING: {e}')
    byclone=defaultdict(list); read_rows=[]
    for r in reads:
        c=align(r,refs[r.sample],minq,trim); byclone[(r.sample,r.clone)].append(c)
        read_rows.append({'sample':r.sample,'clone':r.clone,'primer':r.primer,'file':r.path.name,'read_length':len(r.seq),'mean_quality':round(sum(r.qual)/len(r.qual),2) if r.qual else 0,'orientation':c.orient,'offset':'' if c.offset is None else c.offset,'seed_score':c.seed,'aligned_bases':c.aligned,'high_quality_bases':c.highq,'metadata_result':r.result,'metadata_note':r.note})
    clone_rows=[]; target_rows=[]; problems=[]; need_chrom=set()
    for sample,ref in sorted(refs.items()):
        clones=sorted({c for s,c in byclone if s==sample},key=str) or ['1']
        for clone in clones:
            calls=byclone.get((sample,clone),[]); sel=select(calls); sample_t=targs.get(sample,[])
            tr,ok,miss,bad,last=eval_targets(sample,clone,ref,sel,sample_t); target_rows+=tr
            diffs=[]; bs=base_sets(sel)
            for c in sel: diffs+=c.mismatches+c.indels
            for pos,bases in bs.items():
                if len(bases)>1: diffs.append(prob('conflict',pos,pos-ref.start+1,'','/'.join(sorted(bases)),'','selected reads'))
            for trw in tr:
                if trw['status']!='covered_match': problems.append({'priority':1,'sample':sample,'clone':clone,'final_call':'target_review','kind':'target_'+trw['status'],'map_position':trw['map_positions'],'feature_position':trw['feature_positions'],'expected':trw['expected'],'observed':trw['observed'],'quality':'','read':'; '.join(x.read.path.name for x in sel),'chromatogram_png':''})
            mm=sum(1 for d in diffs if d['kind']=='mismatch'); conf=sum(1 for d in diffs if d['kind']=='conflict'); ind=sum(1 for d in diffs if d['kind'] in ('insertion','deletion')); down=sum(1 for d in diffs if d['kind'] in ('insertion','deletion') and last and d.get('position') and int(d['position'])>last)
            auto,reason,action=clone_call(mode,miss,bad,mm,conf,ind,down,bool(sample_t)); ov=ovs.get((sample,clone),{}); final=ov.get('final_call',auto); mo=ov.get('manual_override',''); mn=ov.get('manual_note',''); reason=mn or reason; action='pick clone' if final=='Candidate' else action
            clone_rows.append({'sample':sample,'clone':clone,'final_call':final,'auto_call':auto,'manual_override':mo,'manual_note':mn,'targets_ok':ok,'targets_missing_or_partial':miss,'targets_problem':bad,'unexpected_mismatch_count':mm,'conflict_count':conf,'indel_review_count':ind,'downstream_indel_count':down,'blocking_indel_count':ind-down if sample_t else ind,'covered_bases':len(bs),'feature_length':ref.end-ref.start+1,'selected_reads':'; '.join(x.read.path.name for x in sel),'all_reads':'; '.join(x.read.path.name for x in calls),'metadata_notes':'; '.join(sorted({x.read.note for x in calls if x.read.note})),'reason':reason,'recommended_action':action})
            sf=[x.read.path.name for x in sel]
            for d in diffs:
                rn=d.get('read',''); files=sf if rn=='selected reads' else [rn]
                problems.append({'priority':{'mismatch':2,'conflict':3,'insertion':4,'deletion':4}.get(d.get('kind',''),9),'sample':sample,'clone':clone,'final_call':final,'kind':d.get('kind',''),'map_position':d.get('position',''),'feature_position':d.get('feature_position',''),'expected':d.get('expected',''),'observed':d.get('observed',''),'quality':d.get('quality',''),'read':'; '.join(x for x in files if x),'chromatogram_png':''})
    for row in problems:
        for rn in str(row.get('read','')).split('; '):
            if rn: need_chrom.add(rn)
    if chro!='none':
        for r in reads:
            if chro=='all' or r.path.name in need_chrom: plot_chrom(r,chrom)
        for row in problems:
            pngs=[]
            for rn in str(row.get('read','')).split('; '):
                p=chrom/(Path(rn).stem+'.png') if rn else None
                if p and p.exists(): pngs.append(str(p))
            row['chromatogram_png']='; '.join(pngs)
    summary=[]
    for sample,ref in sorted(refs.items()):
        rows=[r for r in clone_rows if r['sample']==sample]; cand=[r for r in rows if r['final_call']=='Candidate']
        if cand:
            best=max(cand,key=lambda r:(r['targets_ok'],r['covered_bases'],-r['unexpected_mismatch_count'],-r['indel_review_count'])); sc='Candidate'; act='pick candidate clone(s)'; rea=best['reason']
        elif any(r['final_call']=='Needs review' for r in rows):
            best=max(rows,key=lambda r:(r['targets_ok'],r['covered_bases'])); sc='Needs review'; act='review or resequence best clone'; rea=best['reason']
        else:
            best=rows[0] if rows else {'clone':'','reason':'no reads found'}; sc='No good clone'; act='resequence sample or discard'; rea=best['reason']
        summary.append({'sample':sample,'summary_call':sc,'best_clone':best.get('clone',''),'candidate_clones':'; '.join(f"{sample}-{r['clone']}" for r in sorted(cand,key=lambda r:r['clone'])),'candidate_count':len(cand),'manual_confirmed_clones':'; '.join(f"{sample}-{r['clone']}" for r in sorted(cand,key=lambda r:r['clone']) if r.get('manual_override')),'reference_file':ref.path.name,'feature_range':f'{ref.start}-{ref.end}','clone_calls':'; '.join(f"{r['clone']}:{r['final_call']}" for r in rows),'auto_clone_calls':'; '.join(f"{r['clone']}:{r['auto_call']}" for r in rows),'recommended_action':act,'reason':rea})
    problems.sort(key=lambda r:(r.get('priority',9),str(r.get('sample','')),str(r.get('clone','')),str(r.get('feature_position',''))))
    for r in problems: r.pop('priority',None)
    xlsx=out/f'{prefix}.xlsx'; htm=out/f'{prefix}.html'
    workbook(xlsx,[('Summary',summary),('CloneReview',clone_rows),('TargetSites',target_rows),('ProblemSites',problems),('Reads',read_rows)]); html_report(htm,summary,clone_rows,problems)
    print(f'Wrote {xlsx}'); print(f'Wrote {htm}'); print(f'Samples: {len(summary)}; clones: {len(clone_rows)}; reads: {len(read_rows)}; problems: {len(problems)}')
    return 0
if __name__=='__main__': raise SystemExit(main())
